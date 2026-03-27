import re
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import numpy as np
import pandas as pd

# =========================================================
# CONFIG
# =========================================================
BASE_DIR = Path(".")

FLOOR_DIR = BASE_DIR / "outputs" / "Floor Sheet"
OHLC_DIR = BASE_DIR / "outputs" / "sharesansar"
SECTOR_FILE = BASE_DIR / "outputs" / "Sector" / "sector_master.csv"

FLOOR_PATTERN = "floorsheet_*.csv"
OHLC_PATTERN = "SharePrice_*.csv"

# OUTPUT LOCATION YOU REQUESTED
OUTPUT_DIR = BASE_DIR / "outputs" / "PriceAction"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Excel file name will match the screenshot style, e.g. nepse_signals_2026-03-27.xlsx

DATE_REGEX = re.compile(r"(\d{4}-\d{2}-\d{2})")

# Liquidity filters
MIN_VOLUME_FS = 1000
MIN_NUM_TRADES = 5

# Short-term thresholds
SHORT_STRONG_BUY = 11
SHORT_BUY = 8
SHORT_HOLD = 5

# Long-term thresholds
LONG_STRONG_BUY = 10
LONG_BUY = 7
LONG_HOLD = 4


# =========================================================
# HELPERS
# =========================================================
def extract_date_from_filename(path: Path) -> pd.Timestamp:
    m = DATE_REGEX.search(path.name)
    if not m:
        raise ValueError(f"Could not extract date from filename: {path.name}")
    return pd.to_datetime(m.group(1))


def safe_read_csv(path: Path) -> pd.DataFrame:
    try:
        return pd.read_csv(path)
    except UnicodeDecodeError:
        return pd.read_csv(path, encoding="latin1")


def clean_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(
        series.astype(str).str.replace(",", "", regex=False).str.strip(),
        errors="coerce"
    )


def standardize_symbol(df: pd.DataFrame, col: str = "Symbol") -> pd.DataFrame:
    if col in df.columns:
        df[col] = df[col].astype(str).str.strip().str.upper()
    return df


def safe_div(a, b):
    return np.where((b != 0) & pd.notna(b), a / b, np.nan)


def join_reasons(reasons):
    return " | ".join(reasons) if reasons else "No strong confirmation"


def format_num(x, decimals=2):
    if pd.isna(x):
        return ""
    return f"{x:.{decimals}f}"


def format_pct(x, decimals=2):
    if pd.isna(x):
        return ""
    return f"{x * 100:.{decimals}f}%"


def signal_badge(signal: str) -> str:
    signal = str(signal).strip().upper()
    if signal == "STRONG BUY":
        return '<span class="badge strong-buy">STRONG BUY</span>'
    if signal == "BUY":
        return '<span class="badge buy">BUY</span>'
    if signal == "HOLD":
        return '<span class="badge hold">HOLD</span>'
    if signal == "SELL":
        return '<span class="badge sell">SELL</span>'
    return f'<span class="badge neutral">{signal}</span>'


# =========================================================
# LOAD SHARE PRICE DATA
# ONLY TAKE: Symbol, Open, High, Low, Close
# =========================================================
def load_ohlc_data(ohlc_dir: Path) -> pd.DataFrame:
    files = sorted(ohlc_dir.glob(OHLC_PATTERN))
    if not files:
        raise FileNotFoundError(f"No share price files found in {ohlc_dir}")

    frames = []
    required_cols = {"Symbol", "Open", "High", "Low", "Close"}

    for f in files:
        df = safe_read_csv(f)
        missing = required_cols - set(df.columns)
        if missing:
            continue

        df = df.copy()
        df["Date"] = extract_date_from_filename(f)

        for c in ["Open", "High", "Low", "Close"]:
            df[c] = clean_numeric(df[c])

        df = standardize_symbol(df, "Symbol")
        df = df[["Date", "Symbol", "Open", "High", "Low", "Close"]]
        df = df.dropna(subset=["Date", "Symbol", "Close"])
        frames.append(df)

    if not frames:
        raise ValueError("No valid share price data loaded.")

    out = pd.concat(frames, ignore_index=True)
    out = out.sort_values(["Symbol", "Date"]).drop_duplicates(["Date", "Symbol"], keep="last")
    return out


# =========================================================
# LOAD FLOOR SHEET DATA
# =========================================================
def load_floor_sheet_data(floor_dir: Path) -> pd.DataFrame:
    files = sorted(floor_dir.glob(FLOOR_PATTERN))
    if not files:
        raise FileNotFoundError(f"No floor sheet files found in {floor_dir}")

    frames = []
    required_cols = {"Transact No.", "Symbol", "Buyer", "Seller", "Quantity", "Rate", "Amount"}

    for f in files:
        df = safe_read_csv(f)
        missing = required_cols - set(df.columns)
        if missing:
            continue

        df = df.copy()
        df["Date"] = extract_date_from_filename(f)
        df = standardize_symbol(df, "Symbol")
        df["Buyer"] = df["Buyer"].astype(str).str.strip()
        df["Seller"] = df["Seller"].astype(str).str.strip()

        for c in ["Quantity", "Rate", "Amount"]:
            df[c] = clean_numeric(df[c])

        df = df.dropna(subset=["Date", "Symbol", "Quantity", "Amount"])
        frames.append(df)

    if not frames:
        raise ValueError("No valid floor sheet data loaded.")

    out = pd.concat(frames, ignore_index=True)
    out = out.sort_values(["Date", "Symbol"])
    return out


# =========================================================
# LOAD SECTOR MASTER
# =========================================================
def load_sector_master(sector_file: Path) -> pd.DataFrame:
    if not sector_file.exists():
        return pd.DataFrame(columns=["Symbol", "Sector", "Company"])

    sector = safe_read_csv(sector_file)
    required = {"Symbol", "Sector", "Company"}
    if not required.issubset(sector.columns):
        return pd.DataFrame(columns=["Symbol", "Sector", "Company"])

    sector = standardize_symbol(sector, "Symbol")
    sector = sector[["Symbol", "Sector", "Company"]].drop_duplicates("Symbol")
    return sector


# =========================================================
# FLOOR SHEET AGGREGATION
# VOLUME FROM FLOOR SHEET
# WAP / VWAP FROM FLOOR SHEET
# =========================================================
def compute_broker_concentration(group: pd.DataFrame) -> pd.Series:
    total_qty = group["Quantity"].sum()
    if total_qty <= 0:
        return pd.Series({
            "top_buyer_ratio": 0.0,
            "top_seller_ratio": 0.0,
            "top3_buyer_ratio": 0.0,
            "top3_seller_ratio": 0.0,
            "unique_buyers": 0,
            "unique_sellers": 0,
        })

    buyer_qty = group.groupby("Buyer")["Quantity"].sum().sort_values(ascending=False)
    seller_qty = group.groupby("Seller")["Quantity"].sum().sort_values(ascending=False)

    return pd.Series({
        "top_buyer_ratio": buyer_qty.iloc[0] / total_qty if len(buyer_qty) else 0.0,
        "top_seller_ratio": seller_qty.iloc[0] / total_qty if len(seller_qty) else 0.0,
        "top3_buyer_ratio": buyer_qty.head(3).sum() / total_qty,
        "top3_seller_ratio": seller_qty.head(3).sum() / total_qty,
        "unique_buyers": int(buyer_qty.shape[0]),
        "unique_sellers": int(seller_qty.shape[0]),
    })


def aggregate_floor_sheet_daily(floor: pd.DataFrame) -> pd.DataFrame:
    agg = (
        floor.groupby(["Date", "Symbol"], as_index=False)
        .agg(
            volume_fs=("Quantity", "sum"),
            turnover_fs=("Amount", "sum"),
            num_trades=("Transact No.", "count"),
            avg_rate_fs=("Rate", "mean"),
            max_trade_qty=("Quantity", "max"),
        )
    )

    # day-wise WAP from floor sheet
    agg["wap_fs"] = safe_div(agg["turnover_fs"], agg["volume_fs"])
    agg["avg_trade_size"] = safe_div(agg["volume_fs"], agg["num_trades"])

    broker = floor.groupby(["Date", "Symbol"]).apply(compute_broker_concentration).reset_index()
    out = agg.merge(broker, on=["Date", "Symbol"], how="left")
    out["imbalance"] = out["top_buyer_ratio"] - out["top_seller_ratio"]

    return out


# =========================================================
# MERGE DATA
# =========================================================
def build_master_dataset(ohlc: pd.DataFrame, floor_daily: pd.DataFrame, sector: pd.DataFrame) -> pd.DataFrame:
    df = ohlc.merge(floor_daily, on=["Date", "Symbol"], how="left")
    df = df.merge(sector, on="Symbol", how="left")
    return df.sort_values(["Symbol", "Date"]).reset_index(drop=True)


# =========================================================
# FEATURE ENGINEERING
# =========================================================
def add_features(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    g = df.groupby("Symbol", group_keys=False)

    # Moving averages from Close
    for w in [5, 10, 20, 35, 50]:
        df[f"ma{w}"] = g["Close"].transform(lambda s, w=w: s.rolling(w, min_periods=w).mean())

    # Returns from Close
    for w in [3, 5, 10, 20]:
        df[f"ret_{w}"] = g["Close"].transform(lambda s, w=w: s.pct_change(w))

    # Rolling averages for floor-sheet data
    for w in [10, 20]:
        df[f"vol_fs_avg{w}"] = g["volume_fs"].transform(lambda s, w=w: s.rolling(w, min_periods=w).mean())
        df[f"turnover_fs_avg{w}"] = g["turnover_fs"].transform(lambda s, w=w: s.rolling(w, min_periods=w).mean())
        df[f"num_trades_avg{w}"] = g["num_trades"].transform(lambda s, w=w: s.rolling(w, min_periods=w).mean())
        df[f"imbalance_avg{w}"] = g["imbalance"].transform(lambda s, w=w: s.rolling(w, min_periods=w).mean())

    # Ratios using floor-sheet volume
    df["vol_ratio_10"] = safe_div(df["volume_fs"], df["vol_fs_avg10"])
    df["vol_ratio_20"] = safe_div(df["volume_fs"], df["vol_fs_avg20"])
    df["turnover_ratio_10"] = safe_div(df["turnover_fs"], df["turnover_fs_avg10"])
    df["turnover_ratio_20"] = safe_div(df["turnover_fs"], df["turnover_fs_avg20"])
    df["num_trades_ratio_10"] = safe_div(df["num_trades"], df["num_trades_avg10"])
    df["num_trades_ratio_20"] = safe_div(df["num_trades"], df["num_trades_avg20"])

    # WAP strength from floor-sheet WAP
    df["wap_diff"] = df["Close"] - df["wap_fs"]
    df["wap_strength"] = safe_div(df["wap_diff"], df["wap_fs"])

    # Breakouts
    df["high_5"] = g["High"].transform(lambda s: s.rolling(5, min_periods=5).max())
    df["high_20"] = g["High"].transform(lambda s: s.rolling(20, min_periods=20).max())
    df["prev_high_5"] = g["high_5"].shift(1)
    df["prev_high_20"] = g["high_20"].shift(1)
    df["breakout_5d"] = (df["Close"] > df["prev_high_5"]).astype(int)
    df["breakout_20d"] = (df["Close"] > df["prev_high_20"]).astype(int)

    # Multi-day bullish confirmation
    df["bull_day_short"] = (
        (df["Close"] > df["ma10"]) &
        (df["Close"] > df["wap_fs"]) &
        (df["ret_3"] > 0) &
        (df["imbalance"] > 0)
    ).astype(int)

    df["bull_day_long"] = (
        (df["Close"] > df["ma20"]) &
        (df["Close"] > df["wap_fs"]) &
        (df["ret_10"] > 0) &
        (df["imbalance"] > 0)
    ).astype(int)

    df["confirm_3d_short"] = g["bull_day_short"].transform(lambda s: s.rolling(3, min_periods=3).sum())
    df["confirm_5d_long"] = g["bull_day_long"].transform(lambda s: s.rolling(5, min_periods=5).sum())

    # Persistence features
    df["wap_strength_avg3"] = g["wap_strength"].transform(lambda s: s.rolling(3, min_periods=3).mean())
    df["wap_strength_avg5"] = g["wap_strength"].transform(lambda s: s.rolling(5, min_periods=5).mean())
    df["top_buy_persist_3"] = g["top_buyer_ratio"].transform(lambda s: s.rolling(3, min_periods=3).mean())
    df["top_sell_persist_3"] = g["top_seller_ratio"].transform(lambda s: s.rolling(3, min_periods=3).mean())

    return df


# =========================================================
# SHORT-TERM MODEL
# =========================================================
def short_term_score_and_reason(row):
    score = 0
    reasons = []
    negative_reasons = []

    if pd.notna(row["volume_fs"]) and row["volume_fs"] >= MIN_VOLUME_FS:
        score += 1
        reasons.append("Good liquidity from floor-sheet volume")
    else:
        negative_reasons.append("Low floor-sheet volume")

    if pd.notna(row["num_trades"]) and row["num_trades"] >= MIN_NUM_TRADES:
        score += 1
        reasons.append("Enough trade activity")
    else:
        negative_reasons.append("Low number of trades")

    if pd.notna(row["ma5"]) and row["Close"] > row["ma5"]:
        score += 1
        reasons.append("Close is above MA5")
    else:
        negative_reasons.append("Close is below MA5")

    if pd.notna(row["ma10"]) and row["Close"] > row["ma10"]:
        score += 2
        reasons.append("Close is above MA10")
    else:
        negative_reasons.append("Close is below MA10")

    if pd.notna(row["ret_3"]) and row["ret_3"] > 0:
        score += 1
        reasons.append("3-day momentum is positive")
    else:
        negative_reasons.append("3-day momentum is weak")

    if pd.notna(row["ret_5"]) and row["ret_5"] > 0:
        score += 1
        reasons.append("5-day momentum is positive")
    else:
        negative_reasons.append("5-day momentum is weak")

    if pd.notna(row["vol_ratio_10"]) and row["vol_ratio_10"] > 1.15:
        score += 1
        reasons.append("Floor-sheet volume is above 10-day average")
    else:
        negative_reasons.append("Floor-sheet volume is not strong vs 10-day average")

    if pd.notna(row["turnover_ratio_10"]) and row["turnover_ratio_10"] > 1.10:
        score += 1
        reasons.append("Turnover is above 10-day average")
    else:
        negative_reasons.append("Turnover is not strong vs 10-day average")

    if pd.notna(row["num_trades_ratio_10"]) and row["num_trades_ratio_10"] > 1.05:
        score += 1
        reasons.append("Trade count is increasing")
    else:
        negative_reasons.append("Trade count is not expanding")

    if pd.notna(row["wap_fs"]) and row["Close"] > row["wap_fs"]:
        score += 2
        reasons.append("Close is above floor-sheet WAP, buyers controlled the session")
    else:
        negative_reasons.append("Close is below WAP, sellers had better control")

    if pd.notna(row["imbalance"]) and row["imbalance"] > 0:
        score += 2
        reasons.append("Buyer broker concentration is stronger than seller broker concentration")
    else:
        negative_reasons.append("Broker imbalance is not positive")

    if pd.notna(row["top_buyer_ratio"]) and row["top_buyer_ratio"] > 0.25:
        score += 1
        reasons.append("Top buyer broker concentration is high")
    else:
        negative_reasons.append("Top buyer broker concentration is not strong")

    if pd.notna(row["confirm_3d_short"]) and row["confirm_3d_short"] >= 2:
        score += 2
        reasons.append("At least 2 of the last 3 sessions confirmed bullish conditions")
    else:
        negative_reasons.append("Recent sessions do not confirm strong short-term trend")

    if pd.notna(row["wap_strength_avg3"]) and row["wap_strength_avg3"] > 0:
        score += 1
        reasons.append("3-day WAP strength is positive")
    else:
        negative_reasons.append("3-day WAP strength is weak")

    if int(row.get("breakout_5d", 0)) == 1:
        score += 1
        reasons.append("Price broke above recent 5-day high")
    else:
        negative_reasons.append("No 5-day breakout")

    if (
        pd.notna(row["wap_fs"]) and row["wap_fs"] > row["Close"]
        and pd.notna(row["top_sell_persist_3"]) and row["top_sell_persist_3"] > 0.28
    ):
        signal = "SELL"
        insight = "SELL because price is below WAP and seller-broker pressure has persisted for 3 days"
        why = join_reasons(["Price below WAP", "Persistent seller concentration", "Distribution warning"])
        return score, signal, insight, why

    if score >= SHORT_STRONG_BUY:
        signal = "STRONG BUY"
        insight = "Strong short-term bullish setup with trend, momentum, floor-sheet volume, WAP, and broker confirmation aligned"
        why = join_reasons(reasons)
    elif score >= SHORT_BUY:
        signal = "BUY"
        insight = "Short-term bullish setup with multiple confirmations, but not the strongest possible"
        why = join_reasons(reasons)
    elif score >= SHORT_HOLD:
        signal = "HOLD"
        insight = "Mixed short-term setup. Some bullish signs exist, but confirmation is incomplete"
        why = join_reasons(reasons + negative_reasons[:3])
    else:
        signal = "SELL"
        insight = "Weak short-term structure. Trend, floor-sheet volume, participation, or broker control is not supportive"
        why = join_reasons(negative_reasons[:6])

    return score, signal, insight, why


# =========================================================
# LONG-TERM MODEL
# =========================================================
def long_term_score_and_reason(row):
    score = 0
    reasons = []
    negative_reasons = []

    if pd.notna(row["volume_fs"]) and row["volume_fs"] >= MIN_VOLUME_FS:
        score += 1
        reasons.append("Good liquidity from floor-sheet volume")
    else:
        negative_reasons.append("Low floor-sheet volume")

    if pd.notna(row["num_trades"]) and row["num_trades"] >= MIN_NUM_TRADES:
        score += 1
        reasons.append("Enough trade activity")
    else:
        negative_reasons.append("Low number of trades")

    if pd.notna(row["ma20"]) and row["Close"] > row["ma20"]:
        score += 2
        reasons.append("Close is above MA20")
    else:
        negative_reasons.append("Close is below MA20")

    if pd.notna(row["ma35"]) and row["Close"] > row["ma35"]:
        score += 2
        reasons.append("Close is above MA35")
    else:
        negative_reasons.append("Close is below MA35")

    if pd.notna(row["ma50"]) and row["Close"] > row["ma50"]:
        score += 1
        reasons.append("Close is above MA50")
    else:
        negative_reasons.append("Close is below MA50")

    if pd.notna(row["ret_10"]) and row["ret_10"] > 0:
        score += 2
        reasons.append("10-day momentum is positive")
    else:
        negative_reasons.append("10-day momentum is weak")

    if pd.notna(row["ret_20"]) and row["ret_20"] > 0:
        score += 1
        reasons.append("20-day momentum is positive")
    else:
        negative_reasons.append("20-day momentum is weak")

    if pd.notna(row["vol_ratio_20"]) and row["vol_ratio_20"] > 1.10:
        score += 1
        reasons.append("Floor-sheet volume is above 20-day average")
    else:
        negative_reasons.append("Floor-sheet volume is not strong vs 20-day average")

    if pd.notna(row["turnover_ratio_20"]) and row["turnover_ratio_20"] > 1.05:
        score += 1
        reasons.append("Turnover is above 20-day average")
    else:
        negative_reasons.append("Turnover is not strong vs 20-day average")

    if pd.notna(row["wap_fs"]) and row["Close"] > row["wap_fs"]:
        score += 1
        reasons.append("Close is above floor-sheet WAP")
    else:
        negative_reasons.append("Close is below WAP")

    if pd.notna(row["imbalance_avg20"]) and row["imbalance_avg20"] > 0:
        score += 2
        reasons.append("20-day average broker imbalance is positive")
    else:
        negative_reasons.append("20-day broker imbalance is not positive")

    if (
        pd.notna(row["top3_buyer_ratio"])
        and pd.notna(row["top3_seller_ratio"])
        and row["top3_buyer_ratio"] > row["top3_seller_ratio"]
    ):
        score += 1
        reasons.append("Top 3 buyer brokers are stronger than top 3 seller brokers")
    else:
        negative_reasons.append("Top 3 buyer brokers are not dominating")

    if pd.notna(row["confirm_5d_long"]) and row["confirm_5d_long"] >= 3:
        score += 2
        reasons.append("At least 3 of the last 5 sessions confirmed bullish conditions")
    else:
        negative_reasons.append("Recent sessions do not confirm strong long-term trend")

    if pd.notna(row["wap_strength_avg5"]) and row["wap_strength_avg5"] > 0:
        score += 1
        reasons.append("5-day WAP strength is positive")
    else:
        negative_reasons.append("5-day WAP strength is weak")

    if int(row.get("breakout_20d", 0)) == 1:
        score += 1
        reasons.append("Price broke above recent 20-day high")
    else:
        negative_reasons.append("No 20-day breakout")

    if (
        pd.notna(row["wap_fs"]) and row["wap_fs"] > row["Close"]
        and pd.notna(row["top3_seller_ratio"]) and row["top3_seller_ratio"] > 0.55
    ):
        signal = "SELL"
        insight = "SELL because price is below WAP and top seller brokers dominate the structure"
        why = join_reasons(["Price below WAP", "Top 3 seller brokers dominate", "Distribution warning"])
        return score, signal, insight, why

    if score >= LONG_STRONG_BUY:
        signal = "STRONG BUY"
        insight = "Strong longer-term bullish structure with trend, momentum, floor-sheet volume, WAP, and broker persistence aligned"
        why = join_reasons(reasons)
    elif score >= LONG_BUY:
        signal = "BUY"
        insight = "Longer-term bullish structure is positive, though not at the strongest level"
        why = join_reasons(reasons)
    elif score >= LONG_HOLD:
        signal = "HOLD"
        insight = "Mixed longer-term structure. Trend is not fully broken, but confirmation is incomplete"
        why = join_reasons(reasons + negative_reasons[:3])
    else:
        signal = "SELL"
        insight = "Weak longer-term structure. Trend, floor-sheet volume, broker persistence, or momentum is not supportive"
        why = join_reasons(negative_reasons[:6])

    return score, signal, insight, why


# =========================================================
# APPLY MODELS
# =========================================================
def apply_models(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    short_results = out.apply(short_term_score_and_reason, axis=1)
    out["short_score"] = [x[0] for x in short_results]
    out["short_signal"] = [x[1] for x in short_results]
    out["short_insight"] = [x[2] for x in short_results]
    out["short_why"] = [x[3] for x in short_results]

    long_results = out.apply(long_term_score_and_reason, axis=1)
    out["long_score"] = [x[0] for x in long_results]
    out["long_signal"] = [x[1] for x in long_results]
    out["long_insight"] = [x[2] for x in long_results]
    out["long_why"] = [x[3] for x in long_results]

    return out


# =========================================================
# FINAL SHEET BUILDERS
# =========================================================
def latest_rows(df: pd.DataFrame) -> pd.DataFrame:
    return (
        df.sort_values(["Symbol", "Date"])
        .groupby("Symbol", as_index=False)
        .tail(1)
        .reset_index(drop=True)
    )


def build_short_sheet(df_latest: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "Date", "Symbol", "Sector", "Company", "Open", "High", "Low", "Close",
        "volume_fs", "wap_fs", "num_trades",
        "ma5", "ma10", "ret_3", "ret_5",
        "vol_ratio_10", "turnover_ratio_10",
        "imbalance", "top_buyer_ratio", "top_seller_ratio",
        "confirm_3d_short", "wap_strength_avg3", "breakout_5d",
        "short_score", "short_signal", "short_insight", "short_why"
    ]
    out = df_latest[[c for c in cols if c in df_latest.columns]].copy()
    return out.sort_values(["short_score", "Symbol"], ascending=[False, True])


def build_long_sheet(df_latest: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "Date", "Symbol", "Sector", "Company", "Open", "High", "Low", "Close",
        "volume_fs", "wap_fs", "num_trades",
        "ma20", "ma35", "ma50", "ret_10", "ret_20",
        "vol_ratio_20", "turnover_ratio_20",
        "imbalance_avg20", "top3_buyer_ratio", "top3_seller_ratio",
        "confirm_5d_long", "wap_strength_avg5", "breakout_20d",
        "long_score", "long_signal", "long_insight", "long_why"
    ]
    out = df_latest[[c for c in cols if c in df_latest.columns]].copy()
    return out.sort_values(["long_score", "Symbol"], ascending=[False, True])


# =========================================================
# EXCEL REPORT
# =========================================================
SIGNAL_FILL_MAP = {
    "STRONG BUY": "14532D",
    "BUY": "166534",
    "HOLD": "92400E",
    "SELL": "991B1B",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="334155"),
    right=Side(style="thin", color="334155"),
    top=Side(style="thin", color="334155"),
    bottom=Side(style="thin", color="334155"),
)
TITLE_FILL = PatternFill("solid", fgColor="0F172A")
SECTION_FILL = PatternFill("solid", fgColor="1E293B")
SUMMARY_FILL = PatternFill("solid", fgColor="E2E8F0")

def sanitize_sheet_title(name: str) -> str:
    invalid = set('[]:*?/\\')
    cleaned = ''.join('_' if ch in invalid else ch for ch in name)
    return cleaned[:31]

def format_sheet_values(df: pd.DataFrame, kind: str) -> pd.DataFrame:
    out = df.copy()
    if "Date" in out.columns:
        out["Date"] = pd.to_datetime(out["Date"]).dt.strftime("%Y-%m-%d")

    two_dec_cols = {"Open", "High", "Low", "Close", "wap_fs", "ma5", "ma10", "ma20", "ma35", "ma50", "vol_ratio_10", "vol_ratio_20", "turnover_ratio_10", "turnover_ratio_20", "imbalance", "imbalance_avg20"}
    int_cols = {"volume_fs", "num_trades", "confirm_3d_short", "confirm_5d_long", "breakout_5d", "breakout_20d", "short_score", "long_score"}
    pct_cols = {"ret_3", "ret_5", "ret_10", "ret_20", "top_buyer_ratio", "top_seller_ratio", "top3_buyer_ratio", "top3_seller_ratio", "wap_strength_avg3", "wap_strength_avg5"}

    for col in out.columns:
        if col in two_dec_cols:
            out[col] = out[col].map(format_num)
        elif col in int_cols:
            out[col] = out[col].map(lambda x: format_num(x, 0))
        elif col in pct_cols:
            out[col] = out[col].map(format_pct)

    return out

def autofit_worksheet(ws):
    for col_cells in ws.columns:
        col_idx = col_cells[0].column
        max_len = 0
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            if '\n' in val:
                val = max(val.split('\n'), key=len)
            max_len = max(max_len, len(val))
        width = min(max(max_len + 2, 10), 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def style_header_row(ws, row_idx: int):
    for cell in ws[row_idx]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def apply_body_style(ws, start_row: int, end_row: int, signal_col_name: str | None = None):
    headers = {cell.value: cell.column for cell in ws[start_row - 1] if cell.value is not None}
    signal_col = headers.get(signal_col_name) if signal_col_name else None

    for r in range(start_row, end_row + 1):
        fill_color = "111827" if (r - start_row) % 2 == 0 else "0B1220"
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(color="E5E7EB")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER

        if signal_col:
            signal_cell = ws.cell(r, signal_col)
            signal = str(signal_cell.value).strip().upper()
            if signal in SIGNAL_FILL_MAP:
                signal_cell.fill = PatternFill("solid", fgColor=SIGNAL_FILL_MAP[signal])
                signal_cell.font = Font(color="FFFFFF", bold=True)
                signal_cell.alignment = Alignment(horizontal="center", vertical="center")

def add_table(ws, start_row: int, end_row: int, display_name: str):
    ref = f"A{start_row}:{get_column_letter(ws.max_column)}{end_row}"
    table = Table(displayName=display_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

def write_section_table(ws, start_row: int, title: str, df: pd.DataFrame, table_name: str, signal_column: str):
    last_col = get_column_letter(max(1, len(df.columns)))
    ws.merge_cells(f"A{start_row}:{last_col}{start_row}")
    title_cell = ws.cell(start_row, 1, title)
    title_cell.fill = SECTION_FILL
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    header_row = start_row + 1
    for c_idx, col in enumerate(df.columns, start=1):
        ws.cell(header_row, c_idx, col)
    style_header_row(ws, header_row)

    for r_offset, row in enumerate(df.itertuples(index=False), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(header_row + r_offset, c_idx, value)

    data_start = header_row + 1
    data_end = header_row + len(df)
    if len(df) > 0:
        apply_body_style(ws, data_start, data_end, signal_column)
        add_table(ws, header_row, data_end, table_name)

    return data_end + 2

def create_excel_report(short_df: pd.DataFrame, long_df: pd.DataFrame, output_dir: Path) -> Path:
    latest_date = None
    for frame in (short_df, long_df):
        if not frame.empty and "Date" in frame.columns:
            latest_date = pd.to_datetime(frame["Date"]).max()
            break
    if latest_date is None:
        latest_date = pd.Timestamp(datetime.now().date())

    output_file = output_dir / f"nepse_signals_{latest_date.strftime('%Y-%m-%d')}.xlsx"

    short_fmt = format_sheet_values(short_df, "short")
    long_fmt = format_sheet_values(long_df, "long")

    wb = Workbook()
    ws = wb.active
    ws.title = sanitize_sheet_title(latest_date.strftime("%Y-%m-%d"))
    ws.sheet_view.showGridLines = False

    max_cols = max(len(short_fmt.columns), len(long_fmt.columns), 4)
    last_col = get_column_letter(max_cols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "NEPSE Price Action Report"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"Daily floorsheet + market overview + price action report ({datetime.now().strftime('%I:%M %p')})"
    ws["A2"].fill = TITLE_FILL
    ws["A2"].font = Font(color="CBD5E1", italic=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    short_counts = short_df["short_signal"].value_counts().to_dict() if not short_df.empty else {}
    long_counts = long_df["long_signal"].value_counts().to_dict() if not long_df.empty else {}

    ws["A4"] = "Short-Term Summary"
    ws["A5"] = f"STRONG BUY: {short_counts.get('STRONG BUY', 0)} | BUY: {short_counts.get('BUY', 0)} | HOLD: {short_counts.get('HOLD', 0)} | SELL: {short_counts.get('SELL', 0)}"
    ws["C4"] = "Long-Term Summary"
    ws["C5"] = f"STRONG BUY: {long_counts.get('STRONG BUY', 0)} | BUY: {long_counts.get('BUY', 0)} | HOLD: {long_counts.get('HOLD', 0)} | SELL: {long_counts.get('SELL', 0)}"
    for cell_ref in ("A4", "C4"):
        ws[cell_ref].font = Font(bold=True, color="0F172A")
        ws[cell_ref].fill = SUMMARY_FILL
    for cell_ref in ("A5", "C5"):
        ws[cell_ref].alignment = Alignment(wrap_text=True)

    next_row = 7
    next_row = write_section_table(ws, next_row, "Short-Term Signals", short_fmt, "ShortTermSignals", "short_signal")
    next_row = write_section_table(ws, next_row, "Long-Term Signals", long_fmt, "LongTermSignals", "long_signal")

    ws.freeze_panes = "A8"
    autofit_worksheet(ws)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20

    wb.save(output_file)
    return output_file

def main():
    ohlc = load_ohlc_data(OHLC_DIR)
    floor = load_floor_sheet_data(FLOOR_DIR)
    sector = load_sector_master(SECTOR_FILE)

    floor_daily = aggregate_floor_sheet_daily(floor)
    df = build_master_dataset(ohlc, floor_daily, sector)
    df = add_features(df)
    df = apply_models(df)

    # keep rows with at least enough history for short-term features
    valid_rows = df.groupby("Symbol").cumcount() + 1
    df_model = df.loc[valid_rows >= 20].copy()

    latest = latest_rows(df_model)

    short_sheet = build_short_sheet(latest)
    long_sheet = build_long_sheet(latest)

    output_file = create_excel_report(short_sheet, long_sheet, OUTPUT_DIR)
    print(f"Saved Excel report: {output_file}")


if __name__ == "__main__":
    main()
