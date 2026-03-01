# scripts/generate_trading_report.py
# ------------------------------------------------------------
# Advanced NEPSE Trading Insight Report (Excel) – Retail-Pro Edition
# Data sources (GitHub repo structure):
# - outputs/Floor Sheet/floorsheet_YYYY-MM-DD.csv
# - outputs/sharesansar/SharePrice_YYYY-MM-DD.csv
# - outputs/Sector/sector_master.csv  (Symbol, Company, Sector/Sectors)
# Optional:
# - outputs/Brokers/broker_master.csv (Broker, BrokerName, BrokerType)
#
# Windows:
# - 1D / 7D / 15D / 1M(30 trading days)
#
# Upgrades in this version:
# - Dynamic Top-N for pressure (1D=4, 7D=5, 15D+=10)
# - Volume Surge metrics (latest vs window-average)
# - Volatility & candle strength metrics (range%, close position)
# - Trade Setups sheet (actionable retail-friendly shortlist per window)
# - Risk flags (operator pressure, sell-wall risk, weak follow-through)
# ------------------------------------------------------------

import re
import json
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

# =========================
# CONFIG
# =========================
ROOT = Path(__file__).resolve().parents[1]

FLOOR_DIR = ROOT / "outputs" / "Floor Sheet"
PRICE_DIR = ROOT / "outputs" / "sharesansar"

SECTOR_DIR = ROOT / "outputs" / "Sector"
SECTOR_PATH = (
    SECTOR_DIR / "sector_master.csv"
    if (SECTOR_DIR / "sector_master.csv").exists()
    else (SECTOR_DIR / "sector_master.csv.csv")
)

BROKER_DIR = ROOT / "outputs" / "Brokers"
BROKER_PATH = BROKER_DIR / "broker_master.csv"

REPORT_DIR = ROOT / "outputs" / "reports"
REPORT_DIR.mkdir(parents=True, exist_ok=True)

WINDOWS = {"1D": 1, "7D": 7, "15D": 15, "1M": 30}

FLOOR_RE = re.compile(r".*?(\d{4}-\d{2}-\d{2}).*\.csv$", re.IGNORECASE)
PRICE_RE = re.compile(r".*?(\d{4}-\d{2}-\d{2}).*\.csv$", re.IGNORECASE)

CRORE = 10_000_000


# =========================
# HELPERS
# =========================
def safe_float(v):
    if v is None:
        return np.nan
    try:
        s = str(v).replace(",", "").strip()
        if s == "" or s.lower() == "nan":
            return np.nan
        return float(s)
    except Exception:
        return np.nan


def list_dates_from_folder(folder: Path, regex: re.Pattern):
    dates, files = [], []
    if not folder.exists():
        return dates, files
    for p in folder.glob("*.csv"):
        m = regex.match(p.name)
        if not m:
            continue
        d = m.group(1)
        try:
            dt = pd.to_datetime(d).date()
        except Exception:
            continue
        dates.append(dt)
        files.append(p)
    dates_files = sorted(zip(dates, files), key=lambda x: x[0])
    return [d for d, _ in dates_files], [f for _, f in dates_files]


def choose_window_dates(all_dates_sorted, n):
    if not all_dates_sorted:
        return []
    return all_dates_sorted[-min(n, len(all_dates_sorted)) :]


def window_topn(window_name: str) -> int:
    """Dynamic Top-N sizing suited for NEPSE retail context."""
    if window_name == "1D":
        return 4
    if window_name == "7D":
        return 5
    return 10  # 15D / 1M and longer


def load_sector_master(path: Path):
    if not path.exists():
        return pd.DataFrame(columns=["Symbol", "Company", "Sectors"])
    df = pd.read_csv(path)
    df.columns = [c.strip() for c in df.columns]
    if "Sector" in df.columns and "Sectors" not in df.columns:
        df = df.rename(columns={"Sector": "Sectors"})
    for c in ["Symbol", "Company", "Sectors"]:
        if c not in df.columns:
            df[c] = ""
    df["Symbol"] = df["Symbol"].astype(str).str.strip().str.upper()
    return df[["Symbol", "Company", "Sectors"]].drop_duplicates()


def load_broker_master(path: Path):
    if not path.exists():
        return pd.DataFrame(columns=["Broker", "BrokerName", "BrokerType"])
    df = pd.read_csv(path)
    df.columns = [c.strip() for c in df.columns]
    for c in ["Broker", "BrokerName", "BrokerType"]:
        if c not in df.columns:
            df[c] = "" if c != "Broker" else np.nan
    df["Broker"] = pd.to_numeric(df["Broker"], errors="coerce").astype("Int64")
    df["BrokerName"] = df["BrokerName"].astype(str).str.strip()
    df["BrokerType"] = df["BrokerType"].astype(str).str.strip().str.upper()
    df.loc[~df["BrokerType"].isin(["INSTITUTION", "OPERATOR", "RETAIL", "UNKNOWN"]), "BrokerType"] = "UNKNOWN"
    return df[["Broker", "BrokerName", "BrokerType"]].drop_duplicates()


def read_floorsheet_file(path: Path, trade_date):
    df = pd.read_csv(path)
    df.columns = [c.strip() for c in df.columns]

    ren = {
        "Transact No": "Transact No.",
        "Transact No.": "Transact No.",
        "Transaction No.": "Transact No.",
        "Qty": "Quantity",
        "QTY": "Quantity",
        "Amt": "Amount",
    }
    df = df.rename(columns={k: v for k, v in ren.items() if k in df.columns})

    needed = ["Transact No.", "Symbol", "Buyer", "Seller", "Quantity", "Rate", "Amount"]
    missing = [c for c in needed if c not in df.columns]
    if missing:
        raise ValueError(f"Floorsheet missing columns {missing} in {path.name}")

    df["TradeDate"] = pd.to_datetime(trade_date)
    df["Quantity"] = df["Quantity"].apply(safe_float).fillna(0).astype(float)
    df["Rate"] = df["Rate"].apply(safe_float).fillna(0).astype(float)
    df["Amount"] = df["Amount"].apply(safe_float).fillna(0).astype(float)

    for col in ["Buyer", "Seller"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")

    df["Symbol"] = df["Symbol"].astype(str).str.strip().str.upper()
    return df[needed + ["TradeDate"]]


def read_price_file(path: Path, trade_date):
    df = pd.read_csv(path)
    df.columns = [c.strip() for c in df.columns]
    if "Volume" in df.columns and "Vol" not in df.columns:
        df = df.rename(columns={"Volume": "Vol"})

    core = ["Symbol", "Open", "High", "Low", "Close", "LTP"]
    core_missing = [c for c in core if c not in df.columns]
    if core_missing:
        raise ValueError(f"SharePrice missing columns {core_missing} in {path.name}")

    for c in ["VWAP", "Vol", "Turnover"]:
        if c not in df.columns:
            df[c] = np.nan

    df["TradeDate"] = pd.to_datetime(trade_date)
    for c in ["Open", "High", "Low", "Close", "LTP", "VWAP", "Vol", "Turnover"]:
        df[c] = df[c].apply(safe_float)

    df["Symbol"] = df["Symbol"].astype(str).str.strip().str.upper()
    keep_cols = ["TradeDate", "Symbol", "Open", "High", "Low", "Close", "LTP", "VWAP", "Vol", "Turnover"]
    return df[keep_cols]


def zscore(s: pd.Series):
    s = s.replace([np.inf, -np.inf], np.nan).fillna(0.0)
    std = s.std(ddof=0)
    if std == 0 or np.isnan(std):
        return pd.Series(np.zeros(len(s)), index=s.index)
    return (s - s.mean()) / (std + 1e-9)


# =========================
# METRICS
# =========================
def symbol_metrics_from_floorsheet(fs: pd.DataFrame):
    fs = fs.copy()
    fs["_qxr"] = fs["Quantity"] * fs["Rate"]
    g = fs.groupby("Symbol", as_index=False).agg(
        Trades=("Transact No.", "count"),
        Total_Qty=("Quantity", "sum"),
        Total_Amount=("Amount", "sum"),
        _qxr=("_qxr", "sum"),
    )
    g["VWAP"] = np.where(g["Total_Qty"] > 0, g["_qxr"] / g["Total_Qty"], np.nan)
    g["Total_Amount_Cr"] = g["Total_Amount"] / CRORE
    return g.drop(columns=["_qxr"], errors="ignore")


def broker_symbol_metrics(fs: pd.DataFrame):
    buy = fs[["TradeDate", "Symbol", "Buyer", "Quantity", "Rate"]].copy()
    buy = buy.rename(columns={"Buyer": "Broker"})
    buy["Buy_Qty"] = buy["Quantity"]
    buy["Sell_Qty"] = 0.0
    buy["_buy_cost"] = buy["Quantity"] * buy["Rate"]
    buy["_buy_amt"] = buy["Quantity"] * buy["Rate"]

    sell = fs[["TradeDate", "Symbol", "Seller", "Quantity", "Rate"]].copy()
    sell = sell.rename(columns={"Seller": "Broker"})
    sell["Buy_Qty"] = 0.0
    sell["Sell_Qty"] = sell["Quantity"]
    sell["_buy_cost"] = 0.0
    sell["_buy_amt"] = 0.0

    x = pd.concat([buy, sell], ignore_index=True)
    x["Broker"] = x["Broker"].astype("Int64")

    g = x.groupby(["TradeDate", "Symbol", "Broker"], as_index=False).agg(
        Buy_Qty=("Buy_Qty", "sum"),
        Sell_Qty=("Sell_Qty", "sum"),
        _buy_cost=("_buy_cost", "sum"),
        _buy_amt=("_buy_amt", "sum"),
    )
    g["Net_Qty"] = g["Buy_Qty"] - g["Sell_Qty"]
    g["Avg_Buy_Cost"] = np.where(g["Buy_Qty"] > 0, g["_buy_cost"] / g["Buy_Qty"], np.nan)
    g["Buy_Amount_Cr"] = g["_buy_amt"] / CRORE
    return g.drop(columns=["_buy_cost", "_buy_amt"], errors="ignore")


def top_net_brokers(bsym_symbol_level: pd.DataFrame, topn=5):
    buyers = (
        bsym_symbol_level.sort_values(["Symbol", "Net_Qty"], ascending=[True, False])
        .groupby("Symbol")
        .head(topn)
        .copy()
    )
    sellers = (
        bsym_symbol_level.sort_values(["Symbol", "Net_Qty"], ascending=[True, True])
        .groupby("Symbol")
        .head(topn)
        .copy()
    )
    buyers["Side"] = f"Top{topn}_Net_Buyers"
    sellers["Side"] = f"Top{topn}_Net_Sellers"
    return pd.concat([buyers, sellers], ignore_index=True)


def compute_pressure(bsym_symbol_level: pd.DataFrame, topn: int):
    """
    Buy_Pressure: dominance share of top-N positive net brokers
    Sell_Pressure: dominance share of top-N negative net brokers (abs)
    """
    if bsym_symbol_level.empty:
        return pd.DataFrame(columns=["Symbol", "Buy_Pressure", "Sell_Pressure"])

    t = bsym_symbol_level.copy()
    t["pos"] = t["Net_Qty"].clip(lower=0)
    t["neg_abs"] = (-t["Net_Qty"]).clip(lower=0)

    pos_total = t.groupby("Symbol", as_index=False)["pos"].sum().rename(columns={"pos": "pos_total"})
    neg_total = t.groupby("Symbol", as_index=False)["neg_abs"].sum().rename(columns={"neg_abs": "neg_total"})

    top_pos = (
        t.sort_values(["Symbol", "pos"], ascending=[True, False])
        .groupby("Symbol")
        .head(topn)
        .groupby("Symbol", as_index=False)["pos"]
        .sum()
        .rename(columns={"pos": "top_pos"})
    )
    top_neg = (
        t.sort_values(["Symbol", "neg_abs"], ascending=[True, False])
        .groupby("Symbol")
        .head(topn)
        .groupby("Symbol", as_index=False)["neg_abs"]
        .sum()
        .rename(columns={"neg_abs": "top_neg"})
    )

    out = (
        pos_total.merge(top_pos, on="Symbol", how="left")
        .merge(neg_total, on="Symbol", how="left")
        .merge(top_neg, on="Symbol", how="left")
    )
    out["Buy_Pressure"] = np.where(out["pos_total"] > 0, out["top_pos"].fillna(0) / out["pos_total"], np.nan)
    out["Sell_Pressure"] = np.where(out["neg_total"] > 0, out["top_neg"].fillna(0) / out["neg_total"], np.nan)
    return out[["Symbol", "Buy_Pressure", "Sell_Pressure"]]


def momentum_from_prices(price_window: pd.DataFrame, latest_date: pd.Timestamp):
    if price_window.empty:
        return pd.DataFrame(columns=["Symbol", "Momentum"])
    p = price_window.sort_values(["Symbol", "TradeDate"])
    first = p.groupby("Symbol", as_index=False).first()[["Symbol", "Close"]].rename(columns={"Close": "Close_first"})
    last = p[p["TradeDate"] == latest_date][["Symbol", "Close"]].rename(columns={"Close": "Close_last"})
    out = first.merge(last, on="Symbol", how="inner")
    out["Momentum"] = np.where(out["Close_first"] > 0, (out["Close_last"] / out["Close_first"]) - 1.0, np.nan)
    return out[["Symbol", "Momentum"]]


def volume_surge_from_prices(price_window: pd.DataFrame, latest_date: pd.Timestamp):
    """
    Volume surge = latest Vol / avg Vol in the window (excluding latest day if possible).
    """
    if price_window.empty:
        return pd.DataFrame(columns=["Symbol", "Vol_Surge", "Vol_Avg", "Vol_Latest"])

    p = price_window.copy()
    p["Vol"] = pd.to_numeric(p["Vol"], errors="coerce")
    latest = p[p["TradeDate"] == latest_date][["Symbol", "Vol"]].rename(columns={"Vol": "Vol_Latest"})
    hist = p[p["TradeDate"] != latest_date][["Symbol", "Vol"]]
    if hist.empty:
        # if window is 1 day, use same day as avg to avoid division by 0
        avg = latest.rename(columns={"Vol_Latest": "Vol_Avg"})
    else:
        avg = hist.groupby("Symbol", as_index=False)["Vol"].mean().rename(columns={"Vol": "Vol_Avg"})

    out = latest.merge(avg, on="Symbol", how="left")
    out["Vol_Surge"] = np.where(
        out["Vol_Avg"].notna() & (out["Vol_Avg"] > 0),
        out["Vol_Latest"] / out["Vol_Avg"],
        np.nan,
    )
    return out[["Symbol", "Vol_Surge", "Vol_Avg", "Vol_Latest"]]


def candle_metrics_from_prices(price_window: pd.DataFrame, latest_date: pd.Timestamp):
    """
    Adds simple retail-friendly candle context from latest day:
    - Range_% = (High-Low)/Low * 100
    - Close_Pos = (Close-Low)/(High-Low)  -> near 1 means closed near high
    - Body_% = abs(Close-Open)/Open * 100
    """
    if price_window.empty:
        return pd.DataFrame(columns=["Symbol", "Range_%", "Close_Pos", "Body_%"])

    p = price_window[price_window["TradeDate"] == latest_date].copy()
    if p.empty:
        return pd.DataFrame(columns=["Symbol", "Range_%", "Close_Pos", "Body_%"])

    for c in ["Open", "High", "Low", "Close"]:
        p[c] = pd.to_numeric(p[c], errors="coerce")

    rng = (p["High"] - p["Low"])
    p["Range_%"] = np.where(p["Low"] > 0, (rng / p["Low"]) * 100, np.nan)
    p["Close_Pos"] = np.where(rng > 0, (p["Close"] - p["Low"]) / rng, np.nan)
    p["Body_%"] = np.where(p["Open"] > 0, (p["Close"] - p["Open"]).abs() / p["Open"] * 100, np.nan)
    return p[["Symbol", "Range_%", "Close_Pos", "Body_%"]]


# =========================
# SCORING: BUY/HOLD/SELL
# =========================
def classify_signal(score: pd.Series):
    bins = [-1e9, 50, 70, 1e9]
    labels = ["SELL / AVOID", "HOLD", "BUY"]
    return pd.cut(score, bins=bins, labels=labels, right=False)


def build_trade_score(sym: pd.DataFrame):
    x = sym.copy()

    x["Price_vs_VWAP"] = np.where(x["VWAP"].notna() & (x["VWAP"] > 0), (x["Last_Price"] / x["VWAP"]) - 1.0, np.nan)
    x["Close_gt_VWAP"] = np.where(
        x["VWAP"].notna() & (x["VWAP"] > 0) & x["Last_Price"].notna(),
        (x["Last_Price"] >= x["VWAP"]),
        False,
    )

    x["Activity"] = np.log1p(x["Total_Qty"].fillna(0))
    x["Liq"] = np.log1p(x["Total_Amount_Cr"].fillna(0))

    x["BuyP"] = x["Buy_Pressure"].fillna(0)
    x["SellP"] = x["Sell_Pressure"].fillna(0)

    # Volume surge gives breakout confirmation; cap to reduce outlier blow-ups
    x["Vol_Surge"] = pd.to_numeric(x.get("Vol_Surge", np.nan), errors="coerce")
    x["Vol_Surge_c"] = x["Vol_Surge"].clip(lower=0, upper=10).fillna(1.0)  # 1=normal

    raw = (
        zscore(x["Price_vs_VWAP"].fillna(0)) * 22
        + zscore(x["Momentum"].fillna(0)) * 22
        + zscore(x["Activity"]) * 14
        + zscore(x["Liq"]) * 14
        + (zscore(x["BuyP"]) - zscore(x["SellP"])) * 18
        + zscore(np.log1p(x["Vol_Surge_c"])) * 10
    )

    raw_min, raw_max = float(raw.min()), float(raw.max())
    x["Score"] = np.where(raw_max > raw_min, 100 * (raw - raw_min) / (raw_max - raw_min), 50.0)
    x["Recommendation"] = classify_signal(x["Score"])

    reasons = []
    for _, r in x.iterrows():
        tags = []
        pvv = float(r.get("Price_vs_VWAP", 0) or 0)
        mom = float(r.get("Momentum", 0) or 0)
        bp = float(r.get("Buy_Pressure", 0) or 0)
        sp = float(r.get("Sell_Pressure", 0) or 0)
        liq = float(r.get("Total_Amount_Cr", 0) or 0)
        vs = float(r.get("Vol_Surge", 1) or 1)

        if pvv > 0.01:
            tags.append("Above VWAP")
        elif pvv < -0.01:
            tags.append("Below VWAP")

        if mom > 0.03:
            tags.append("Strong Momentum")
        elif mom < -0.03:
            tags.append("Weak Momentum")

        if bp >= 0.40:
            tags.append("Buyer Dominance")
        if sp >= 0.40:
            tags.append("Seller Dominance")

        if vs >= 1.8:
            tags.append("Volume Surge")

        if liq >= 5:
            tags.append("High Liquidity")

        reasons.append(", ".join(tags[:5]))
    x["Reason"] = reasons
    return x


# =========================
# SMART MONEY / INSTITUTION / OPERATOR
# =========================
def build_smart_money(symbol_level_broker: pd.DataFrame, latest_date: pd.Timestamp, scored_symbols: pd.DataFrame):
    if symbol_level_broker.empty:
        sm_sym = pd.DataFrame(
            columns=[
                "Symbol",
                "Net_Qty",
                "Buy_Qty",
                "Sell_Qty",
                "Active_Days",
                "Top3_Pos_Share",
                "SmartMoneyScore",
                "SmartMoneySignal",
            ]
        )
        sm_broker = pd.DataFrame(
            columns=[
                "Broker",
                "Net_Qty",
                "Buy_Qty",
                "Sell_Qty",
                "Active_Days",
                "Symbols",
                "SmartBrokerScore",
                "Tag",
            ]
        )
        return sm_sym, sm_broker

    bs = symbol_level_broker.copy()

    if "TradeDate" in bs.columns:
        daily = bs.copy()
        ad = daily.groupby(["Symbol", "Broker"], as_index=False).agg(
            Active_Days=("Net_Qty", lambda s: int((s > 0).sum()))
        )
        bs_w = daily.groupby(["Symbol", "Broker"], as_index=False).agg(
            Buy_Qty=("Buy_Qty", "sum"),
            Sell_Qty=("Sell_Qty", "sum"),
            Net_Qty=("Net_Qty", "sum"),
            Avg_Buy_Cost=("Avg_Buy_Cost", "mean"),
            Buy_Amount_Cr=("Buy_Amount_Cr", "sum"),
        )
        bs = bs_w.merge(ad, on=["Symbol", "Broker"], how="left")
    else:
        if "Active_Days" not in bs.columns:
            bs["Active_Days"] = np.nan

    sym_tot = bs.groupby("Symbol", as_index=False).agg(
        Buy_Qty=("Buy_Qty", "sum"),
        Sell_Qty=("Sell_Qty", "sum"),
        Net_Qty=("Net_Qty", "sum"),
        Net_Buy_Amount_Cr=("Buy_Amount_Cr", "sum"),
        Accum_Broker_Count=("Net_Qty", lambda s: int((s > 0).sum())),
        Active_Days=("Active_Days", lambda s: int(np.nanmax(s)) if np.any(~np.isnan(s)) else np.nan),
    )

    tmp = bs.copy()
    tmp["pos"] = tmp["Net_Qty"].clip(lower=0)
    pos_sum = tmp.groupby("Symbol", as_index=False)["pos"].sum().rename(columns={"pos": "pos_total"})
    top3 = (
        tmp.sort_values(["Symbol", "pos"], ascending=[True, False])
        .groupby("Symbol")
        .head(3)
        .groupby("Symbol", as_index=False)["pos"]
        .sum()
        .rename(columns={"pos": "top3_pos"})
    )
    dom = pos_sum.merge(top3, on="Symbol", how="left")
    dom["Top3_Pos_Share"] = np.where(dom["pos_total"] > 0, dom["top3_pos"].fillna(0) / dom["pos_total"], np.nan)
    dom = dom[["Symbol", "Top3_Pos_Share"]]

    sm_sym = sym_tot.merge(dom, on="Symbol", how="left")

    cols = [
        "Symbol",
        "Company",
        "Sectors",
        "VWAP",
        "Last_Price",
        "Momentum",
        "Buy_Pressure",
        "Sell_Pressure",
        "Total_Amount_Cr",
        "Vol_Surge",
        "Range_%",
        "Close_Pos",
    ]
    ctx = scored_symbols[[c for c in cols if c in scored_symbols.columns]].drop_duplicates("Symbol")
    sm_sym = sm_sym.merge(ctx, on="Symbol", how="left")

    sm_sym["Price_vs_VWAP_pct"] = np.where(
        sm_sym["VWAP"].notna() & (sm_sym["VWAP"] > 0),
        (sm_sym["Last_Price"] / sm_sym["VWAP"] - 1) * 100,
        np.nan,
    )

    base = (
        zscore(sm_sym["Net_Qty"]) * 33
        + zscore(sm_sym["Net_Buy_Amount_Cr"].fillna(0)) * 18
        + zscore(sm_sym["Active_Days"].fillna(0)) * 14
        + zscore(sm_sym["Top3_Pos_Share"].fillna(0)) * 13
        + zscore(sm_sym["Price_vs_VWAP_pct"].fillna(0) / 100.0) * 12
        + zscore(np.log1p(pd.to_numeric(sm_sym["Vol_Surge"], errors="coerce").fillna(1.0))) * 10
    )
    mn, mx = float(base.min()), float(base.max())
    sm_sym["SmartMoneyScore"] = np.where(mx > mn, 100 * (base - mn) / (mx - mn), 50.0)

    def sm_signal(v):
        if pd.isna(v):
            return "NEUTRAL"
        if v >= 75:
            return "🟢 ACCUMULATION"
        if v >= 55:
            return "🟡 EARLY"
        if v < 35:
            return "🔴 DISTRIBUTION"
        return "⚪ NEUTRAL"

    sm_sym["SmartMoneySignal"] = sm_sym["SmartMoneyScore"].apply(sm_signal)

    sm_broker = bs.groupby("Broker", as_index=False).agg(
        Buy_Qty=("Buy_Qty", "sum"),
        Sell_Qty=("Sell_Qty", "sum"),
        Net_Qty=("Net_Qty", "sum"),
        Active_Days=("Active_Days", lambda s: int(np.nanmax(s)) if np.any(~np.isnan(s)) else np.nan),
        Symbols=("Symbol", "nunique"),
        Buy_Amount_Cr=("Buy_Amount_Cr", "sum"),
    )
    braw = (
        zscore(sm_broker["Net_Qty"]) * 45
        + zscore(sm_broker["Buy_Amount_Cr"].fillna(0)) * 25
        + zscore(sm_broker["Symbols"].fillna(0)) * 15
        + zscore(sm_broker["Active_Days"].fillna(0)) * 15
    )
    bmn, bmx = float(braw.min()), float(braw.max())
    sm_broker["SmartBrokerScore"] = np.where(bmx > bmn, 100 * (braw - bmn) / (bmx - bmn), 50.0)
    sm_broker["Tag"] = np.where(
        sm_broker["SmartBrokerScore"] >= 75,
        "SMART-BROKER",
        np.where(sm_broker["SmartBrokerScore"] >= 55, "WATCH", "NORMAL"),
    )
    return sm_sym, sm_broker


def build_institution_operator(
    bs_window: pd.DataFrame, price_latest: pd.DataFrame, sector: pd.DataFrame, broker_master: pd.DataFrame
):
    if bs_window.empty:
        inst = pd.DataFrame(
            columns=[
                "Broker",
                "BrokerName",
                "BrokerType",
                "Net_Qty",
                "Buy_Qty",
                "Sell_Qty",
                "Buy_Amount_Cr",
                "Active_Days",
                "Symbols",
                "Top_Sector",
                "Concentration_Pct",
                "Flip_Ratio",
                "InstitutionScore",
                "Tag",
            ]
        )
        opr = pd.DataFrame(
            columns=[
                "Broker",
                "BrokerName",
                "BrokerType",
                "Symbol",
                "Buy_Qty",
                "Sell_Qty",
                "Net_Qty",
                "Buy_Amount_Cr",
                "Active_Days",
                "Flip_Ratio",
                "Concentration_Pct",
                "Avg_Buy_Cost",
                "Last_Price",
                "Cost_vs_LTP_pct",
                "OperatorScore",
                "Tag",
            ]
        )
        return inst, opr

    b = bs_window.copy()
    for c in ["Buy_Qty", "Sell_Qty", "Net_Qty", "Buy_Amount_Cr", "Avg_Buy_Cost", "Active_Days"]:
        if c not in b.columns:
            b[c] = np.nan

    bro_tot = b.groupby("Broker", as_index=False).agg(
        Buy_Qty=("Buy_Qty", "sum"),
        Sell_Qty=("Sell_Qty", "sum"),
        Net_Qty=("Net_Qty", "sum"),
        Buy_Amount_Cr=("Buy_Amount_Cr", "sum"),
        Active_Days=("Active_Days", lambda s: int(np.nanmax(s)) if np.any(~np.isnan(s)) else np.nan),
        Symbols=("Symbol", "nunique"),
    )

    bro_tot["Flip_Ratio"] = np.where(
        (bro_tot["Buy_Qty"] + bro_tot["Sell_Qty"]) > 0,
        np.minimum(bro_tot["Buy_Qty"], bro_tot["Sell_Qty"])
        / np.maximum(bro_tot["Buy_Qty"], bro_tot["Sell_Qty"]).replace(0, np.nan),
        np.nan,
    )

    tmp = b.copy()
    tmp["abs_net"] = tmp["Net_Qty"].abs()
    den = tmp.groupby("Broker", as_index=False)["abs_net"].sum().rename(columns={"abs_net": "abs_net_total"})
    num = tmp.groupby("Broker", as_index=False)["abs_net"].max().rename(columns={"abs_net": "abs_net_max"})
    conc = den.merge(num, on="Broker", how="left")
    conc["Concentration_Pct"] = np.where(
        conc["abs_net_total"] > 0, (conc["abs_net_max"] / conc["abs_net_total"]) * 100, np.nan
    )
    bro_tot = bro_tot.merge(conc[["Broker", "Concentration_Pct"]], on="Broker", how="left")

    b2 = b.merge(sector[["Symbol", "Sectors"]], on="Symbol", how="left")
    sec_focus = (
        b2.groupby(["Broker", "Sectors"], as_index=False)["Net_Qty"]
        .sum()
        .assign(abs_net=lambda d: d["Net_Qty"].abs())
        .sort_values(["Broker", "abs_net"], ascending=[True, False])
        .groupby("Broker")
        .head(1)
        .rename(columns={"Sectors": "Top_Sector"})
    )
    bro_tot = bro_tot.merge(sec_focus[["Broker", "Top_Sector"]], on="Broker", how="left")

    inst = bro_tot.merge(broker_master, on="Broker", how="left")
    inst["BrokerName"] = inst["BrokerName"].fillna("")
    inst["BrokerType"] = inst["BrokerType"].fillna("UNKNOWN")

    raw_inst = (
        zscore(inst["Net_Qty"]) * 35
        + zscore(inst["Buy_Amount_Cr"].fillna(0)) * 25
        + zscore(inst["Active_Days"].fillna(0)) * 20
        + zscore(inst["Symbols"].fillna(0)) * 15
        - zscore(inst["Flip_Ratio"].fillna(0)) * 15
        - zscore(inst["Concentration_Pct"].fillna(0) / 100.0) * 10
    )
    mn, mx = float(raw_inst.min()), float(raw_inst.max())
    inst["InstitutionScore"] = np.where(mx > mn, 100 * (raw_inst - mn) / (mx - mn), 50.0)
    inst["Tag"] = np.where(inst["InstitutionScore"] >= 75, "INSTITUTION-LIKE", np.where(inst["InstitutionScore"] >= 55, "WATCH", "NORMAL"))

    p = (
        price_latest[["Symbol", "Last_Price"]].drop_duplicates("Symbol")
        if not price_latest.empty
        else pd.DataFrame(columns=["Symbol", "Last_Price"])
    )
    op = b.merge(p, on="Symbol", how="left")

    tot_abs = tmp.groupby("Broker", as_index=False)["abs_net"].sum().rename(columns={"abs_net": "abs_net_total"})
    op = op.merge(tot_abs, on="Broker", how="left")
    op["Concentration_Pct"] = np.where(op["abs_net_total"] > 0, (op["Net_Qty"].abs() / op["abs_net_total"]) * 100, np.nan)

    op["Flip_Ratio"] = np.where(
        (op["Buy_Qty"] + op["Sell_Qty"]) > 0,
        np.minimum(op["Buy_Qty"], op["Sell_Qty"])
        / np.maximum(op["Buy_Qty"], op["Sell_Qty"]).replace(0, np.nan),
        np.nan,
    )
    op["Cost_vs_LTP_pct"] = np.where(
         op["Avg_Buy_Cost"].notna()
         & (op["Avg_Buy_Cost"] > 0)
         & op["Last_Price"].notna()
         & (op["Last_Price"] > 0),
         ((op["Last_Price"] - op["Avg_Buy_Cost"]) / op["Avg_Buy_Cost"]) * 100,
         np.nan
    )
    op = op.merge(broker_master, on="Broker", how="left")
    op["BrokerName"] = op["BrokerName"].fillna("")
    op["BrokerType"] = op["BrokerType"].fillna("UNKNOWN")

    raw_op = (
        zscore(op["Concentration_Pct"].fillna(0) / 100.0) * 35
        + zscore(op["Flip_Ratio"].fillna(0)) * 20
        + zscore(op["Buy_Amount_Cr"].fillna(0)) * 15
        + zscore(op["Net_Qty"].abs().fillna(0)) * 20
        + zscore(op["Cost_vs_LTP_pct"].fillna(0) / 100.0) * 10
    )
    omin, omax = float(raw_op.min()), float(raw_op.max())
    op["OperatorScore"] = np.where(omax > omin, 100 * (raw_op - omin) / (omax - omin), 50.0)
    op["Tag"] = np.where(op["OperatorScore"] >= 75, "OPERATOR-LIKELY", np.where(op["OperatorScore"] >= 55, "WATCH", "NORMAL"))

    op = op.sort_values(["OperatorScore"], ascending=False).head(250).copy()

    inst_cols = [
        "Broker",
        "BrokerName",
        "BrokerType",
        "Net_Qty",
        "Buy_Qty",
        "Sell_Qty",
        "Buy_Amount_Cr",
        "Active_Days",
        "Symbols",
        "Top_Sector",
        "Concentration_Pct",
        "Flip_Ratio",
        "InstitutionScore",
        "Tag",
    ]
    op_cols = [
        "Broker",
        "BrokerName",
        "BrokerType",
        "Symbol",
        "Buy_Qty",
        "Sell_Qty",
        "Net_Qty",
        "Buy_Amount_Cr",
        "Active_Days",
        "Flip_Ratio",
        "Concentration_Pct",
        "Avg_Buy_Cost",
        "Last_Price",
        "Cost_vs_LTP_pct",
        "OperatorScore",
        "Tag",
    ]

    inst = inst[[c for c in inst_cols if c in inst.columns]].copy()
    op = op[[c for c in op_cols if c in op.columns]].copy()
    return inst, op


# =========================
# EXCEL FORMATTING
# =========================
def style_sheet(ws):
    ws.freeze_panes = "A2"
    thin = Side(style="thin", color="D0D7DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="EEF2FF")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border = border
            if isinstance(c.value, str) and len(c.value) > 60:
                c.alignment = Alignment(wrap_text=True, vertical="top")


def add_table(ws, df, name="Table1"):
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    style_sheet(ws)

    nrows = ws.max_row
    ncols = ws.max_column
    if nrows >= 2 and ncols >= 1:
        from openpyxl.utils import get_column_letter

        ref = f"A1:{get_column_letter(ncols)}{nrows}"
        tab = Table(displayName=name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)


def find_col(ws, header_name: str):
    for idx, cell in enumerate(ws[1], start=1):
        if str(cell.value).strip() == header_name:
            return idx
    return None


def apply_recommendation_cf(ws, rec_col_name="Recommendation"):
    col = find_col(ws, rec_col_name)
    if not col or ws.max_row < 2:
        return
    from openpyxl.utils import get_column_letter

    cl = get_column_letter(col)
    rng = f"{cl}2:{cl}{ws.max_row}"

    fill_green = PatternFill("solid", fgColor="C6EFCE")
    fill_red = PatternFill("solid", fgColor="FFC7CE")
    fill_yellow = PatternFill("solid", fgColor="FFEB9C")

    first_cell = f"{cl}2"
    ws.conditional_formatting.add(
        rng, FormulaRule(formula=[f'UPPER({first_cell})="BUY"'], fill=fill_green, stopIfTrue=True)
    )
    ws.conditional_formatting.add(
        rng, FormulaRule(formula=[f'UPPER({first_cell})="SELL / AVOID"'], fill=fill_red, stopIfTrue=True)
    )
    ws.conditional_formatting.add(
        rng, FormulaRule(formula=[f'UPPER({first_cell})="HOLD"'], fill=fill_yellow, stopIfTrue=True)
    )


def apply_score_scale(ws, score_col_name="Score"):
    col = find_col(ws, score_col_name)
    if not col or ws.max_row < 2:
        return
    from openpyxl.utils import get_column_letter

    cl = get_column_letter(col)
    rng = f"{cl}2:{cl}{ws.max_row}"
    rule = ColorScaleRule(
        start_type="min",
        start_value=0,
        start_color="F8696B",
        mid_type="percentile",
        mid_value=50,
        mid_color="FFEB84",
        end_type="max",
        end_value=100,
        end_color="63BE7B",
    )
    ws.conditional_formatting.add(rng, rule)


def auto_fit_columns(ws, max_width=45, sample_limit=200):
    for col in ws.columns:
        max_len = 10
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is None:
                continue
            text = str(cell.value)
            max_len = max(max_len, len(text[:sample_limit]))
        ws.column_dimensions[col_letter].width = min(max_width, max_len + 2)


def add_bar_chart(ws, data_col_header: str, category_col_header: str, title: str, top_n=20, anchor="J2"):
    data_col = find_col(ws, data_col_header)
    cat_col = find_col(ws, category_col_header)
    if not data_col or not cat_col or ws.max_row < 3:
        return

    max_row = min(ws.max_row, top_n + 1)
    data = Reference(ws, min_col=data_col, min_row=1, max_row=max_row)
    cats = Reference(ws, min_col=cat_col, min_row=2, max_row=max_row)

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = data_col_header
    chart.x_axis.title = category_col_header
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = False

    ws.add_chart(chart, anchor)


def write_excel_report(path: Path, sheets: dict, charts_config: list):
    wb = Workbook()
    wb.remove(wb.active)

    for i, (sname, df) in enumerate(sheets.items(), start=1):
        ws = wb.create_sheet(title=sname[:31])
        df = df.replace([np.inf, -np.inf], np.nan)
        add_table(ws, df, name=f"T{i}")

        apply_recommendation_cf(ws, "Recommendation")
        apply_score_scale(ws, "Score")
        apply_score_scale(ws, "SmartMoneyScore")
        apply_score_scale(ws, "InstitutionScore")
        apply_score_scale(ws, "OperatorScore")
        apply_score_scale(ws, "SmartBrokerScore")
        apply_score_scale(ws, "SetupScore")

        auto_fit_columns(ws)

    for cfg in charts_config:
        sname = cfg.get("sheet")
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        add_bar_chart(
            ws,
            data_col_header=cfg["data_col"],
            category_col_header=cfg["cat_col"],
            title=cfg["title"],
            top_n=cfg.get("top_n", 20),
            anchor=cfg.get("anchor", "J2"),
        )

    wb.save(path)


# =========================
# MAIN
# =========================
def main():
    floor_dates, floor_files = list_dates_from_folder(FLOOR_DIR, FLOOR_RE)
    price_dates, price_files = list_dates_from_folder(PRICE_DIR, PRICE_RE)

    if not floor_dates:
        raise RuntimeError(f"No floorsheet csv files found in {FLOOR_DIR}")

    trading_dates = floor_dates
    latest_date = pd.to_datetime(trading_dates[-1])

    sector = load_sector_master(SECTOR_PATH)
    brokers_master = load_broker_master(BROKER_PATH)

    print("✅ Floorsheet files:", len(floor_files))
    print("✅ Share price files:", len(price_files))
    print("✅ Sector path:", SECTOR_PATH, "exists:", SECTOR_PATH.exists(), "rows:", len(sector))
    print("✅ Broker master path:", BROKER_PATH, "exists:", BROKER_PATH.exists(), "rows:", len(brokers_master))
    print("✅ Latest trading date:", trading_dates[-1])

    floor_map = {d: f for d, f in zip(floor_dates, floor_files)}
    price_map = {d: f for d, f in zip(price_dates, price_files)}

    symbol_summary_all, top_picks_all = [], []
    broker_summary_all, broker_by_symbol_all = [], []
    market_overview_rows, price_movers_all = [], []
    sector_summary_all = []
    smart_money_all, smart_broker_all = [], []
    inst_all, opr_all = [], []
    setups_all = []

    for wname, n in WINDOWS.items():
        w_dates = choose_window_dates(trading_dates, n)
        if not w_dates:
            continue
        w_latest = pd.to_datetime(w_dates[-1])
        topn = window_topn(wname)

        # Load floorsheets
        fs_list = []
        for d in w_dates:
            fp = floor_map.get(d)
            if fp is None:
                continue
            fs_list.append(read_floorsheet_file(fp, d))
        fs = pd.concat(fs_list, ignore_index=True) if fs_list else pd.DataFrame()

        # Load prices
        pr_list = []
        for d in w_dates:
            pp = price_map.get(d)
            if pp is None:
                continue
            pr_list.append(read_price_file(pp, d))
        pr = pd.concat(pr_list, ignore_index=True) if pr_list else pd.DataFrame()

        # Latest price snapshot
        price_latest = (
            pr[pr["TradeDate"] == w_latest][["Symbol", "Close", "LTP", "VWAP", "Vol", "Turnover", "Open", "High", "Low"]]
            .rename(columns={"LTP": "Last_Price", "Close": "Close_latest"})
            if not pr.empty
            else pd.DataFrame(
                columns=["Symbol", "Close_latest", "Last_Price", "VWAP", "Vol", "Turnover", "Open", "High", "Low"]
            )
        )

        # Floorsheet metrics
        sym = symbol_metrics_from_floorsheet(fs) if not fs.empty else pd.DataFrame(
            columns=["Symbol", "Trades", "Total_Qty", "Total_Amount", "VWAP", "Total_Amount_Cr"]
        )
        sym = sym.merge(price_latest[["Symbol", "Last_Price", "Vol", "Turnover"]], on="Symbol", how="left")

        # Broker-symbol daily metrics
        bs_daily = broker_symbol_metrics(fs) if not fs.empty else pd.DataFrame(
            columns=["TradeDate", "Symbol", "Broker", "Buy_Qty", "Sell_Qty", "Net_Qty", "Avg_Buy_Cost", "Buy_Amount_Cr"]
        )

        # Window-level broker-symbol
        if not bs_daily.empty:
            bs_window = bs_daily.groupby(["Symbol", "Broker"], as_index=False).agg(
                Buy_Qty=("Buy_Qty", "sum"),
                Sell_Qty=("Sell_Qty", "sum"),
                Net_Qty=("Net_Qty", "sum"),
                Avg_Buy_Cost=("Avg_Buy_Cost", "mean"),
                Buy_Amount_Cr=("Buy_Amount_Cr", "sum"),
                Active_Days=("Net_Qty", lambda s: int((s > 0).sum())),
            )
        else:
            bs_window = pd.DataFrame(
                columns=["Symbol", "Broker", "Buy_Qty", "Sell_Qty", "Net_Qty", "Avg_Buy_Cost", "Buy_Amount_Cr", "Active_Days"]
            )

        # Pressure (dynamic top-N)
        pressure = compute_pressure(bs_window, topn=topn)

        # Momentum + Volume surge + Candle metrics
        mom = momentum_from_prices(pr, w_latest)
        vol_surge = volume_surge_from_prices(pr, w_latest)
        candle = candle_metrics_from_prices(pr, w_latest)

        sym = sym.merge(pressure, on="Symbol", how="left").merge(mom, on="Symbol", how="left")
        sym = sym.merge(vol_surge, on="Symbol", how="left").merge(candle, on="Symbol", how="left")
        sym = sym.merge(sector, on="Symbol", how="left")

        # Score
        scored = build_trade_score(sym)

        # Risk flags (simple and actionable)
        scored["Risk_Flags"] = ""
        scored.loc[pd.to_numeric(scored.get("Sell_Pressure", 0), errors="coerce").fillna(0) >= 0.45, "Risk_Flags"] += "SellWall; "
        scored.loc[pd.to_numeric(scored.get("Vol_Surge", 1), errors="coerce").fillna(1) >= 3.0, "Risk_Flags"] += "TooHotVol; "
        scored.loc[pd.to_numeric(scored.get("Close_Pos", 0.5), errors="coerce").fillna(0.5) < 0.35, "Risk_Flags"] += "WeakClose; "

        # Neat columns
        keep_cols = [
            "Symbol", "Company", "Sectors",
            "Trades", "Total_Qty", "Total_Amount_Cr",
            "VWAP", "Last_Price",
            "Buy_Pressure", "Sell_Pressure",
            "Momentum",
            "Vol_Surge", "Range_%", "Close_Pos", "Body_%",
            "Close_gt_VWAP",
            "Score", "Recommendation", "Reason", "Risk_Flags"
        ]
        scored = scored[[c for c in keep_cols if c in scored.columns]].copy()

        # rounding
        for c in ["VWAP", "Last_Price"]:
            if c in scored.columns:
                scored[c] = pd.to_numeric(scored[c], errors="coerce").round(2)
        for c in ["Total_Amount_Cr"]:
            if c in scored.columns:
                scored[c] = pd.to_numeric(scored[c], errors="coerce").round(3)
        for c in ["Buy_Pressure", "Sell_Pressure", "Close_Pos"]:
            if c in scored.columns:
                scored[c] = pd.to_numeric(scored[c], errors="coerce").round(3)
        for c in ["Momentum", "Score", "Vol_Surge", "Range_%", "Body_%"]:
            if c in scored.columns:
                scored[c] = pd.to_numeric(scored[c], errors="coerce").round(2)

        scored.insert(0, "Window", wname)
        symbol_summary_all.append(scored)

        # Top picks
        tp = scored.copy()
        top_buy = tp.sort_values("Score", ascending=False).head(25).assign(List="TOP_BUY")
        top_sell = tp.sort_values("Score", ascending=True).head(25).assign(List="TOP_SELL")
        top_hold = tp[tp["Recommendation"] == "HOLD"].sort_values("Score", ascending=False).head(25).assign(List="TOP_HOLD")
        top_picks_all.append(pd.concat([top_buy, top_hold, top_sell], ignore_index=True))

        # Broker summary
        if not bs_window.empty:
            bsum = bs_window.groupby("Broker", as_index=False).agg(
                Buy_Qty=("Buy_Qty", "sum"),
                Sell_Qty=("Sell_Qty", "sum"),
                Net_Qty=("Net_Qty", "sum"),
                Buy_Amount_Cr=("Buy_Amount_Cr", "sum"),
                Symbols=("Symbol", "nunique"),
            )
            bsum = bsum.merge(brokers_master, on="Broker", how="left")
            bsum["BrokerName"] = bsum["BrokerName"].fillna("")
            bsum["BrokerType"] = bsum["BrokerType"].fillna("UNKNOWN")
            bsum.insert(0, "Window", wname)
            broker_summary_all.append(bsum)

            tb = top_net_brokers(bs_window, topn=topn)
            tb = tb.merge(brokers_master, on="Broker", how="left")
            tb["BrokerName"] = tb["BrokerName"].fillna("")
            tb["BrokerType"] = tb["BrokerType"].fillna("UNKNOWN")

            # ✅ Only for Broker_by_Symbol sheet: add Sector after Symbol
            tb = tb.merge(sector[["Symbol", "Sectors"]], on="Symbol", how="left")
            tb = tb.rename(columns={"Sectors": "Sector"})

            # put Sector right after Symbol
            cols = list(tb.columns)
            if "Symbol" in cols and "Sector" in cols:
                cols.remove("Symbol")
                cols.remove("Sector")
                tb = tb[["Symbol", "Sector"] + cols]

            tb.insert(0, "Window", wname)
            broker_by_symbol_all.append(tb)

        # Market overview
        mrow = {
            "Window": wname,
            "From": str(w_dates[0]),
            "To": str(w_dates[-1]),
            "TopN_Pressure": topn,
            "Symbols_Traded": int(scored["Symbol"].nunique()) if not scored.empty else 0,
            "Total_Amount_Cr": float(pd.to_numeric(scored.get("Total_Amount_Cr", pd.Series([0])), errors="coerce").sum()),
            "Total_Qty": float(pd.to_numeric(scored.get("Total_Qty", pd.Series([0])), errors="coerce").sum()),
            "BUY_Count": int((scored["Recommendation"] == "BUY").sum()) if "Recommendation" in scored.columns else 0,
            "HOLD_Count": int((scored["Recommendation"] == "HOLD").sum()) if "Recommendation" in scored.columns else 0,
            "SELL_Count": int((scored["Recommendation"] == "SELL / AVOID").sum()) if "Recommendation" in scored.columns else 0,
            "Median_Score": float(pd.to_numeric(scored.get("Score", pd.Series([np.nan])), errors="coerce").median(skipna=True)) if not scored.empty else np.nan,
            "Avg_Vol_Surge": float(pd.to_numeric(scored.get("Vol_Surge", pd.Series([np.nan])), errors="coerce").mean(skipna=True)) if not scored.empty else np.nan,
        }
        market_overview_rows.append(mrow)

        # Price movers
        if not pr.empty and pr["TradeDate"].nunique() >= 2:
            p = pr.sort_values(["Symbol", "TradeDate"])
            first = p.groupby("Symbol", as_index=False).first()[["Symbol", "Close"]].rename(columns={"Close": "Close_start"})
            last = p[p["TradeDate"] == w_latest][["Symbol", "Close"]].rename(columns={"Close": "Close_end"})
            mv = first.merge(last, on="Symbol", how="inner")
            mv["Change_%"] = np.where(mv["Close_start"] > 0, (mv["Close_end"] / mv["Close_start"] - 1) * 100, np.nan)
            mv = mv.merge(sector, on="Symbol", how="left")
            mv.insert(0, "Window", wname)
            price_movers_all.append(mv.sort_values("Change_%", ascending=False).head(60))

        # Sector summary
        if not scored.empty and "Sectors" in scored.columns:
            sec = scored.groupby(["Window", "Sectors"], as_index=False).agg(
                Symbols=("Symbol", "nunique"),
                Amount_Cr=("Total_Amount_Cr", "sum"),
                Avg_Score=("Score", "mean"),
                Avg_Momentum=("Momentum", "mean") if "Momentum" in scored.columns else ("Score", "mean"),
                Avg_Vol_Surge=("Vol_Surge", "mean") if "Vol_Surge" in scored.columns else ("Score", "mean"),
            )
            sector_summary_all.append(sec.sort_values(["Window", "Amount_Cr"], ascending=[True, False]))

        # Smart money + Inst/Operator
        sm_sym, sm_broker = build_smart_money(bs_daily, w_latest, scored_symbols=scored)
        sm_sym.insert(0, "Window", wname)
        sm_broker.insert(0, "Window", wname)
        smart_money_all.append(sm_sym)
        smart_broker_all.append(sm_broker)

        inst, opr = build_institution_operator(
            bs_window,
            price_latest[["Symbol", "Last_Price"]] if not price_latest.empty else pd.DataFrame(),
            sector,
            brokers_master,
        )
        inst.insert(0, "Window", wname)
        opr.insert(0, "Window", wname)
        inst_all.append(inst)
        opr_all.append(opr)

        # Trade setups (retail actionable shortlist)
        if not scored.empty:
            sm_ctx = (
                sm_sym[["Symbol", "SmartMoneyScore", "SmartMoneySignal", "Net_Qty", "Net_Buy_Amount_Cr"]]
                if not sm_sym.empty
                else pd.DataFrame()
            )
            op_ctx = (
                opr.groupby("Symbol", as_index=False)["OperatorScore"].max()
                if not opr.empty and "OperatorScore" in opr.columns
                else pd.DataFrame(columns=["Symbol", "OperatorScore"])
            )
            sx = scored.merge(sm_ctx, on="Symbol", how="left").merge(op_ctx, on="Symbol", how="left")
            sx["OperatorScore"] = pd.to_numeric(sx.get("OperatorScore", np.nan), errors="coerce")
            sx["SetupScore"] = (
                pd.to_numeric(sx["Score"], errors="coerce").fillna(50) * 0.55
                + pd.to_numeric(sx.get("SmartMoneyScore", np.nan), errors="coerce").fillna(50) * 0.35
                + zscore(np.log1p(pd.to_numeric(sx.get("Vol_Surge", 1), errors="coerce").fillna(1))) * 10
                - zscore(pd.to_numeric(sx.get("OperatorScore", 0), errors="coerce").fillna(0) / 100.0) * 8
            )
            # normalize to 0-100
            mn, mx = float(sx["SetupScore"].min()), float(sx["SetupScore"].max())
            sx["SetupScore"] = np.where(mx > mn, 100 * (sx["SetupScore"] - mn) / (mx - mn), 50.0)

            # Setup rules (keep simple but strong)
            sx["Setup_Tag"] = "WATCH"
            sx.loc[
                (pd.to_numeric(sx["Score"], errors="coerce") >= 70)
                & (pd.to_numeric(sx.get("Vol_Surge", 1), errors="coerce") >= 1.5)
                & (pd.to_numeric(sx.get("Buy_Pressure", 0), errors="coerce") > pd.to_numeric(sx.get("Sell_Pressure", 0), errors="coerce"))
                & (sx.get("Close_gt_VWAP", False) == True),
                "Setup_Tag",
            ] = "BUY_SETUP"

            sx.loc[
                (pd.to_numeric(sx.get("Sell_Pressure", 0), errors="coerce") >= 0.50)
                | (pd.to_numeric(sx.get("OperatorScore", 0), errors="coerce") >= 75),
                "Setup_Tag",
            ] = "CAUTION"

            # Retail plan hints (not financial advice – generic)
            sx["Retail_Plan"] = ""
            sx.loc[sx["Setup_Tag"] == "BUY_SETUP", "Retail_Plan"] = "Prefer entry near VWAP/Support; cut if closes below VWAP"
            sx.loc[sx["Setup_Tag"] == "CAUTION", "Retail_Plan"] = "Wait for confirmation; avoid chasing; watch sell wall"

            keep = [
                "Window", "Symbol", "Company", "Sectors",
                "SetupScore", "Setup_Tag",
                "Score", "Recommendation", "SmartMoneyScore", "SmartMoneySignal",
                "Vol_Surge", "Momentum", "Buy_Pressure", "Sell_Pressure",
                "Range_%", "Close_Pos", "Risk_Flags",
                "OperatorScore",
                "Retail_Plan",
            ]
            sx = sx[[c for c in keep if c in sx.columns]].copy()
            for c in ["SetupScore", "Score", "SmartMoneyScore", "Vol_Surge", "Momentum", "Range_%", "Close_Pos", "Buy_Pressure", "Sell_Pressure", "OperatorScore"]:
                if c in sx.columns:
                    sx[c] = pd.to_numeric(sx[c], errors="coerce").round(2)
            sx = sx.sort_values(["SetupScore"], ascending=False).head(60).copy()
            setups_all.append(sx)

    # Combine sheets
    symbol_summary = pd.concat(symbol_summary_all, ignore_index=True) if symbol_summary_all else pd.DataFrame()
    top_picks = pd.concat(top_picks_all, ignore_index=True) if top_picks_all else pd.DataFrame()
    broker_summary = pd.concat(broker_summary_all, ignore_index=True) if broker_summary_all else pd.DataFrame()
    broker_by_symbol = pd.concat(broker_by_symbol_all, ignore_index=True) if broker_by_symbol_all else pd.DataFrame()
    market_overview = pd.DataFrame(market_overview_rows)
    price_movers = pd.concat(price_movers_all, ignore_index=True) if price_movers_all else pd.DataFrame()
    sector_summary = pd.concat(sector_summary_all, ignore_index=True) if sector_summary_all else pd.DataFrame()
    smart_money = pd.concat(smart_money_all, ignore_index=True) if smart_money_all else pd.DataFrame()
    smart_broker = pd.concat(smart_broker_all, ignore_index=True) if smart_broker_all else pd.DataFrame()
    inst_tracker = pd.concat(inst_all, ignore_index=True) if inst_all else pd.DataFrame()
    operator_radar = pd.concat(opr_all, ignore_index=True) if opr_all else pd.DataFrame()
    trade_setups = pd.concat(setups_all, ignore_index=True) if setups_all else pd.DataFrame()

    # Chart-friendly sheets (7D default)
    sm_chart = pd.DataFrame()
    sb_chart = pd.DataFrame()
    inst_chart = pd.DataFrame()
    opr_chart = pd.DataFrame()
    setup_chart = pd.DataFrame()

    if not smart_money.empty:
        sm7 = smart_money[smart_money["Window"] == "7D"].copy()
        if not sm7.empty and "SmartMoneyScore" in sm7.columns:
            cols = ["Symbol", "SmartMoneyScore", "SmartMoneySignal", "Net_Qty", "Net_Buy_Amount_Cr", "Sectors"]
            sm_chart = sm7.sort_values("SmartMoneyScore", ascending=False).head(20)[[c for c in cols if c in sm7.columns]].copy()

    if not smart_broker.empty:
        sb7 = smart_broker[smart_broker["Window"] == "7D"].copy()
        if not sb7.empty and "SmartBrokerScore" in sb7.columns:
            sb_chart = sb7.sort_values("SmartBrokerScore", ascending=False).head(15)[
                ["Broker", "SmartBrokerScore", "Net_Qty", "Buy_Amount_Cr", "Tag"]
            ].copy()

    if not inst_tracker.empty:
        inst7 = inst_tracker[inst_tracker["Window"] == "7D"].copy()
        if not inst7.empty and "InstitutionScore" in inst7.columns:
            cols = ["Broker", "InstitutionScore", "Net_Qty", "Buy_Amount_Cr", "Tag", "Top_Sector"]
            inst_chart = inst7.sort_values("InstitutionScore", ascending=False).head(15)[[c for c in cols if c in inst7.columns]].copy()

    if not operator_radar.empty:
        opr7 = operator_radar[operator_radar["Window"] == "7D"].copy()
        if not opr7.empty and "OperatorScore" in opr7.columns:
            opr_chart = opr7.sort_values("OperatorScore", ascending=False).head(20)[
                ["Symbol", "OperatorScore", "Concentration_Pct", "Flip_Ratio", "Tag"]
            ].copy()

    if not trade_setups.empty:
        s7 = trade_setups[trade_setups["Window"] == "7D"].copy()
        if not s7.empty and "SetupScore" in s7.columns:
            setup_chart = s7.sort_values("SetupScore", ascending=False).head(20)[
                ["Symbol", "SetupScore", "Setup_Tag", "Vol_Surge", "Momentum"]
                if "Momentum" in s7.columns
                else ["Symbol", "SetupScore", "Setup_Tag", "Vol_Surge"]
            ].copy()

    readme = pd.DataFrame(
        [
            ["Retail-Pro NEPSE Report", ""],
            ["Windows", "1D=1 trading day, 7D=7 trading days, 15D=15 trading days, 1M=30 trading days (based on available files)"],
            ["Pressure Top-N", "1D=4, 7D=5, 15D/1M=10 (used in Buy/Sell pressure dominance)"],
            ["BUY/HOLD/SELL", "Score (0-100): BUY>=70, HOLD 50-69, SELL/AVOID<50"],
            ["Volume Surge", "Vol_Surge = latest volume / window-average volume (higher helps confirm breakout)"],
            ["Candle Metrics", "Range_% (volatility), Close_Pos (close near high=strong), Body_% (move strength)"],
            ["Trade_Setups", "Shortlist for retail: Score + SmartMoney + Volume Surge, with caution flags for sell-wall/operator pressure"],
            ["Smart Money", "Uses broker net flows + capital + persistence + concentration + price behavior to score accumulation/distribution"],
            ["Institution Tracker", "Behavior-based: persistence + breadth + net flow + capital; penalizes flip & high concentration"],
            ["Operator Radar", "Behavior-based: high concentration + flip + burst + chasing; shows likely operator pressure per broker-symbol"],
            ["Sector file", f"Loaded from: {SECTOR_PATH.name} (exists={SECTOR_PATH.exists()})"],
            ["Broker master", f"Optional: {BROKER_PATH.as_posix()} (exists={BROKER_PATH.exists()}). Columns: Broker,BrokerName,BrokerType"],
        ],
        columns=["Item", "Explanation"],
    )

    sheets = {
        "README": readme,
        "Market_Overview": market_overview.sort_values("Window"),
        "Trade_Setups": trade_setups,
        "Top_Picks": top_picks,
        "Symbol_Summary": symbol_summary,
        "Smart_Money": smart_money,
        "Smart_Brokers": smart_broker,
        "Institution_Tracker": inst_tracker,
        "Operator_Radar": operator_radar,
        "Broker_Summary": broker_summary,
        "Broker_by_Symbol": broker_by_symbol,
        "Sector_Summary": sector_summary,
        "Price_Movers": price_movers,
        "SM_Top20_7D": sm_chart,
        "SB_Top15_7D": sb_chart,
        "Inst_Top15_7D": inst_chart,
        "Opr_Top20_7D": opr_chart,
        "Setup_Top20_7D": setup_chart,
    }

    charts_config = [
        {"sheet": "SM_Top20_7D", "data_col": "SmartMoneyScore", "cat_col": "Symbol", "title": "Top 20 Smart Money Score (7D)", "top_n": 20, "anchor": "H2"},
        {"sheet": "SB_Top15_7D", "data_col": "SmartBrokerScore", "cat_col": "Broker", "title": "Top 15 Smart Brokers (7D)", "top_n": 15, "anchor": "H2"},
        {"sheet": "Inst_Top15_7D", "data_col": "InstitutionScore", "cat_col": "Broker", "title": "Top 15 Institution-like Brokers (7D)", "top_n": 15, "anchor": "H2"},
        {"sheet": "Opr_Top20_7D", "data_col": "OperatorScore", "cat_col": "Symbol", "title": "Top 20 Operator Pressure (7D) by Symbol", "top_n": 20, "anchor": "H2"},
        {"sheet": "Setup_Top20_7D", "data_col": "SetupScore", "cat_col": "Symbol", "title": "Top 20 Trade Setups (7D)", "top_n": 20, "anchor": "H2"},
    ]

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = REPORT_DIR / f"RetailPro_Trading_Insight_Report_{ts}.xlsx"

    write_excel_report(out_path, sheets, charts_config)
    print(f"✅ Report generated: {out_path}")

    latest_json = REPORT_DIR / "latest_report.json"
    latest_json.write_text(json.dumps({"latest_report": out_path.name}, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
